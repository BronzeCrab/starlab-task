from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import permissions, viewsets, views
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response

from books.models import Book, Author
from books.serializers import BookSerializer, AuthorSerializer

from io import BytesIO


class BookViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows books to be viewed or edited.
    """

    queryset = Book.objects.all()
    serializer_class = BookSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ("title", "date_published", "genre")


class AuthorViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows authors to be viewed or edited.
    """

    queryset = Author.objects.all()
    serializer_class = AuthorSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ("name",)


class FileUploadView(views.APIView):
    parser_classes = [FileUploadParser]

    def _parse_worksheet(self, ws: Worksheet, alist: list[str]) -> None:
        """Parsing two worksheets of inpute file."""
        row_ind = 2
        while True:
            cell = ws.cell(row=row_ind, column=2)
            if not cell.value:
                break
            alist.append(str(cell.value).strip().lower())
            row_ind += 1

    def _check_and_update_books(
        self, book_names: list[str], book_authors: list[str]
    ) -> None:
        """Trying to find the book in db and set is_denied to True."""
        if book_names and book_authors:
            for i, book_name in enumerate(book_names):
                book_qs = Book.objects.filter(title__iexact=book_name)
                if book_qs:
                    book = book_qs.first()
                    author_qs = book.author_set.filter(name__iexact=book_authors[i])
                    if author_qs:
                        author = author_qs.first()
                        book.is_denied = True
                        book.save()

    def put(self, request, filename, format=None) -> Response:
        try:
            file_obj = request.data["file"]

            wb = load_workbook(filename=BytesIO(file_obj.read()))

            book_names = []
            ws_name = wb["name"]
            self._parse_worksheet(ws_name, book_names)

            book_authors = []
            ws_name = wb["author"]
            self._parse_worksheet(ws_name, book_authors)

            self._check_and_update_books(book_names, book_authors)

        except Exception as exc:
            print(f"Error: parsing file error: {exc}")
            return Response(status=400)

        return Response(status=204)
